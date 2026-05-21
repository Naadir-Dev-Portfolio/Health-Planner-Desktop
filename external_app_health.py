import json
import os
import subprocess
import sys
import webbrowser
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from PyQt6.QtCore import Qt, QUrl, QSize, QRect, QPoint, QPointF, QTimer
from PyQt6.QtGui import QAction, QPainter, QLinearGradient, QColor, QPen, QIcon, QPixmap
from PyQt6.QtWidgets import (
    QApplication, QWidget, QFrame, QLabel, QPushButton, QVBoxLayout, QHBoxLayout,
    QLineEdit, QScrollArea, QLayout, QSizeGrip, QToolButton, QMenu, QStackedWidget
)
from PyQt6.QtWebEngineWidgets import QWebEngineView
from PyQt6.QtWebEngineCore import QWebEngineProfile, QWebEngineScript, QWebEngineSettings, QWebEnginePage

from openpyxl import load_workbook


# =========================
# APP CONFIG
# =========================
APP_TITLE = "Health Planner Routine App"
WINDOW_OPACITY = 0.93        # 0-1 (1 = Fully opaque)
START_SIZE = (1400, 900)     # primary take workout app dimensions
RADIUS = 10

STATE_DIR_NAME = "widget_positions"

# Paths (relative to this script's directory — standalone mode)
WORKOUT_SUBPATH = Path(".")          # workout_html_files/ sits next to this script
WORKOUT_HTML_DIR = "workout_html_files"
WORKOUT_EXCEL = "master_plan.xlsm"  # Provide your own Excel file — see README

REGIMEN_SUBPATH = Path("regimen_plan")
REGIMEN_INDEX = "index.html"
REGIMEN_EXCEL = "master_plan.xlsm"  # Provide your own Excel file — see README

# Logos — place your own images here or leave blank (app runs without them)
HEALTH_ICON_REL = Path("assets/health_logo.png")
ALT_ICON_REL = Path("assets/logo.png")

# Short labels for alternative workout fields
ALT_FIELD_MAP = {
    "Office Alternative (if commuting)": "OFFICE",
    "Move Alternative (if stress)":      "MOVE",
    "Physical Job Alternative":          "PHYSICAL",
    "Crisis Alternative":                "CRISIS",
}
# =========================


SCROLLBAR_STYLE = """
QScrollBar:vertical { border:none; background:transparent; width:10px; margin:0; }
QScrollBar::handle:vertical { background:#555; min-height:20px; border-radius:5px; }
QScrollBar::handle:vertical:hover { background:#777; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height:0; }
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background:none; }

QScrollBar:horizontal { border:none; background:transparent; height:10px; margin:0; }
QScrollBar::handle:horizontal { background:#555; min-width:20px; border-radius:5px; }
QScrollBar::handle:horizontal:hover { background:#777; }
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width:0; }
QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal { background:none; }
"""

WEB_SCROLLBAR_CSS = """
:root{ color-scheme: dark; }
html, body{ background:#121212 !important; }
::-webkit-scrollbar{ width:10px; height:10px; background:transparent; }
::-webkit-scrollbar-track{ background: rgba(0,0,0,0.35); border-radius: 5px; }
::-webkit-scrollbar-thumb{ background:#555; border-radius:5px; min-height:20px; min-width:20px; }
::-webkit-scrollbar-thumb:hover{ background:#777; }
::-webkit-scrollbar-corner{ background: rgba(0,0,0,0.35); }
"""


def build_theme_qss():
    return f"""
#Root {{ background: transparent; }}

#MainContainer {{
    border: 1px solid #2a2a2a;
    border-radius: {RADIUS}px;
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #2f2f2f,
        stop:0.45 #262626,
        stop:1 #1f1f1f
    );
}}

QFrame#Header {{
    border-top-left-radius: {RADIUS}px;
    border-top-right-radius: {RADIUS}px;
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #3a3a3a,
        stop:0.55 #2f2f2f,
        stop:1 #242424
    );
    border-bottom: 1px solid #2a2a2a;
}}

QLabel#Title {{
    font-size: 20px;
    font-weight: 900;
    color: #f3f3f3;
    letter-spacing: 1px;
}}

QLabel#PhaseLabel {{
    font-size: 11px;
    font-weight: 800;
    color: #7a7a7a;
    letter-spacing: 1px;
    padding-left: 4px;
}}

QPushButton {{
    color:#ffffff;
    font-size: 12px;
    font-weight: 800;
    border: 1px solid #4a4a4a;
    border-radius: 10px;
    padding: 10px 14px;
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #4b4b4b,
        stop:0.55 #3b3b3b,
        stop:1 #2f2f2f
    );
}}
QPushButton:hover {{
    border-color: #5a5a5a;
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #5a5a5a,
        stop:0.55 #474747,
        stop:1 #353535
    );
}}
QPushButton:pressed {{
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #343434,
        stop:1 #262626
    );
}}

QPushButton#NavBtn {{
    font-size: 11px;
    font-weight: 900;
    padding: 8px 12px;
    border-radius: 10px;
}}
QPushButton#NavBtn[active="true"] {{
    color: #111111;
    border-color: #c8c8c8;
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #ffffff,
        stop:1 #e9e9e9
    );
}}
QPushButton#NavBtn[active="false"] {{
    color: #e8e8e8;
    border-color: rgba(255,255,255,0.10);
    background: rgba(0,0,0,0.10);
}}

QPushButton#TodayPrimaryBtn {{
    font-size: 11px;
    font-weight: 900;
    padding: 8px 12px;
    border-radius: 10px;
}}
QPushButton#TodayAltBtn {{
    font-size: 11px;
    font-weight: 800;
    color: #bdbdbd;
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 10px;
    padding: 8px 12px;
    background: rgba(0,0,0,0.10);
}}
QPushButton#TodayAltBtn:hover {{
    color: #ffffff;
    border-color: #4a4a4a;
    background: rgba(255,255,255,0.08);
}}

QLabel#TodaySep {{
    color: #3a3a3a;
    font-size: 14px;
}}

QLabel#TodayNone {{
    font-size: 11px;
    font-weight: 700;
    color: #444444;
    font-style: italic;
    padding-left: 6px;
}}

QFrame#SidePanel {{
    border-radius: {RADIUS}px;
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #1f1f1f,
        stop:1 #171717
    );
    border: 1px solid #2a2a2a;
}}

QScrollArea#WorkoutScroll {{
    background: #121212;
    border: 1px solid #2a2a2a;
    border-radius: {RADIUS}px;
}}
QScrollArea#WorkoutScroll QWidget {{ background: transparent; }}

QLineEdit {{
    background-color:#222222;
    border:1px solid #3a3a3a;
    border-radius: 10px;
    padding: 8px 10px;
    color:#f0f0f0;
}}
QLineEdit:focus {{ border: 1px solid #6b6b6b; }}

QPushButton#YTBtn {{
    color: #ff2b2b;
    border-color: #ff2b2b;
}}
QPushButton#YTBtn:hover {{
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #6a2b2b,
        stop:1 #3a1e1e
    );
}}

QPushButton#YTMusicBtn {{
    color: #ff2b2b;
    border-color: #ff2b2b;
}}
QPushButton#YTMusicBtn:hover {{
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #6a2b2b,
        stop:1 #3a1e1e
    );
}}

QPushButton#MusicBtn {{
    color: #1DB954;
    border-color: #1DB954;
}}
QPushButton#MusicBtn:hover {{
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #2b5a3a,
        stop:1 #1a2f22
    );
}}

QPushButton#MinButton {{
    font: 900 14px 'Segoe UI';
    color: #ffffff;
    border: 1px solid #3f3f3f;
    min-width: 28px; min-height: 28px;
    max-width: 28px; max-height: 28px;
    border-radius: 8px;
    padding: 0px;
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #3b3b3b,
        stop:1 #242424
    );
}}
QPushButton#MinButton:hover {{
    border: 1px solid #5a5a5a;
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #525252,
        stop:1 #303030
    );
}}
QPushButton#MinButton:pressed {{ background: #202020; }}

QPushButton#CloseButton {{
    font: 900 14px 'Segoe UI';
    color: #ffffff;
    border: 1px solid #5a2f2f;
    min-width: 28px; min-height: 28px;
    max-width: 28px; max-height: 28px;
    border-radius: 8px;
    padding: 0px;
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #3a2424,
        stop:1 #261717
    );
}}
QPushButton#CloseButton:hover {{
    border: 1px solid #8a3a3a;
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #6a2f2f,
        stop:1 #3a1e1e
    );
}}
QPushButton#CloseButton:pressed {{ background: #231414; }}

QToolButton#PinBtn {{
    border: 1px solid #2f2f2f;
    border-radius: 10px;
    padding: 0px;
    min-width: 24px;
    min-height: 24px;
    max-width: 24px;
    max-height: 24px;
    color: #d6d6d6;
    background: rgba(0,0,0,0.18);
}}
QToolButton#PinBtn:hover {{
    border-color: #3a3a3a;
    background: rgba(255,255,255,0.08);
}}

{SCROLLBAR_STYLE}

QMenu {{
    background-color: #2b2b2b;
    border: 1px solid #3a3a3a;
    border-radius: 10px;
    padding: 6px;
    color: #f0f0f0;
}}
QMenu::item {{ padding: 6px 12px; border-radius: 6px; }}
QMenu::item:selected {{ background-color: #3a3a3a; }}
QMenu::separator {{ height:1px; background:#3a3a3a; margin:6px 4px; }}

#ErrorOverlay {{
    background: rgba(0,0,0,0.78);
    border-radius: {RADIUS}px;
    color: #ffffff;
    font-size: 13px;
    font-weight: 800;
    padding: 16px;
}}
"""


class UiStateStore:
    def __init__(self):
        script_dir = Path(__file__).resolve().parent
        self.dir = script_dir / STATE_DIR_NAME
        self.dir.mkdir(parents=True, exist_ok=True)
        self.path = self.dir / f"{Path(__file__).stem}.json"

    def load(self) -> dict:
        if not self.path.exists():
            return {}
        try:
            return json.loads(self.path.read_text(encoding="utf-8"))
        except Exception:
            return {}

    def save(self, data: dict) -> None:
        tmp = self.path.with_suffix(".tmp")
        tmp.write_text(json.dumps(data, indent=2), encoding="utf-8")
        tmp.replace(self.path)


class FlowLayout(QLayout):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.itemList = []

    def addItem(self, item):
        self.itemList.append(item)

    def addWidget(self, w):
        super().addWidget(w)

    def count(self):
        return len(self.itemList)

    def itemAt(self, index):
        return self.itemList[index] if 0 <= index < len(self.itemList) else None

    def takeAt(self, index):
        return self.itemList.pop(index) if 0 <= index < len(self.itemList) else None

    def expandingDirections(self):
        return Qt.Orientation(0)

    def hasHeightForWidth(self):
        return True

    def heightForWidth(self, width):
        return self._doLayout(QRect(0, 0, width, 0), True)

    def setGeometry(self, rect):
        super().setGeometry(rect)
        self._doLayout(rect, False)

    def sizeHint(self):
        return self.minimumSize()

    def minimumSize(self):
        res = QSize()
        for i in self.itemList:
            res = res.expandedTo(i.minimumSize())
        return res

    def _doLayout(self, rect, testOnly):
        x, y, line_h = rect.x(), rect.y(), 0
        spacing = 10
        for item in self.itemList:
            next_x = x + item.sizeHint().width() + spacing
            if next_x > rect.right() and line_h > 0:
                x = rect.x()
                y += line_h + spacing
                line_h = 0
            if not testOnly:
                item.setGeometry(QRect(QPoint(x, y), item.sizeHint()))
            x += item.sizeHint().width() + spacing
            line_h = max(line_h, item.sizeHint().height())
        return y + line_h - rect.y()


class WorkoutCard(QFrame):
    def __init__(self, title: str, filename: str, pinned: bool, on_open, on_pin_toggle):
        super().__init__()
        self._filename = filename
        self._on_open = on_open
        self._on_pin_toggle = on_pin_toggle
        self._fav = False

        self.setObjectName("Card")
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setAutoFillBackground(False)
        self.setCursor(Qt.CursorShape.PointingHandCursor)

        self.title_lbl = QLabel(title.upper())
        self.title_lbl.setObjectName("CardTitle")
        self.title_lbl.setStyleSheet("font-size:18px; font-weight:900;")

        self.pin_btn = QToolButton()
        self.pin_btn.setObjectName("PinBtn")
        self.pin_btn.setCheckable(True)
        self.pin_btn.clicked.connect(self._pin_clicked)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(14, 10, 10, 10)
        lay.setSpacing(10)
        lay.addWidget(self.title_lbl, stretch=1)
        lay.addWidget(self.pin_btn)

        self.set_pinned(bool(pinned))

    def _pin_clicked(self):
        self._on_pin_toggle(self._filename)

    def set_pinned(self, pinned: bool):
        self._fav = bool(pinned)

        self.pin_btn.blockSignals(True)
        self.pin_btn.setChecked(self._fav)
        self.pin_btn.setText("\u2605" if self._fav else "\u2606")
        self.pin_btn.blockSignals(False)

        self.setFixedWidth(300)
        self.setFixedHeight(74 if self._fav else 62)

        self.title_lbl.setStyleSheet(
            "color:#111111; font-weight:900; font-size:18px; background:transparent;"
            if self._fav else
            "color:#f0f0f0; font-weight:900; font-size:19px; background:transparent;"
        )

        if self._fav:
            self.pin_btn.setStyleSheet(
                "QToolButton{color:#111111; border:1px solid #c8c8c8; border-radius:10px; "
                "min-width:24px; min-height:24px; max-width:24px; max-height:24px; "
                "background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #ffffff, stop:1 #eeeeee);} "
                "QToolButton:hover{background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #ffffff, stop:1 #f2f2f2);}"
            )
        else:
            self.pin_btn.setStyleSheet("")

        self.update()

    def paintEvent(self, e):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing, True)

        r = self.rect().adjusted(1, 1, -1, -1)
        radius = 10

        pen = QPen(QColor("#d6d6d6" if self._fav else "#2f2f2f"))
        pen.setWidth(1)
        p.setPen(pen)

        grad = QLinearGradient(QPointF(r.left(), r.top()), QPointF(r.left(), r.bottom()))
        if self._fav:
            grad.setColorAt(0.0, QColor("#ffffff"))
            grad.setColorAt(1.0, QColor("#e9e9e9"))
        else:
            grad.setColorAt(0.0, QColor("#1f1f1f"))
            grad.setColorAt(1.0, QColor("#151515"))

        p.setBrush(grad)
        p.drawRoundedRect(r, radius, radius)

        super().paintEvent(e)

    def mousePressEvent(self, e):
        if e.button() == Qt.MouseButton.LeftButton:
            if self.pin_btn.underMouse():
                return
            self._on_open()


class ExternalizingPage(QWebEnginePage):
    def createWindow(self, _type):
        page = QWebEnginePage(self.profile(), self)

        def _handle_url(url: QUrl):
            if url.isValid():
                webbrowser.open(url.toString())

        page.urlChanged.connect(_handle_url)
        return page

    def acceptNavigationRequest(self, url: QUrl, nav_type, is_main_frame: bool):
        if url.isValid():
            u = url.toString()
            if u.startswith("http://") or u.startswith("https://"):
                webbrowser.open(u)
                return False
            if url.isLocalFile():
                return True
        return super().acceptNavigationRequest(url, nav_type, is_main_frame)


@dataclass
class WorkoutPaths:
    base: Path
    workout_dir: Path
    excel_path: Path


@dataclass
class RegimenPaths:
    site_dir: Path
    html_path: Path
    excel_path: Path


class WorkoutView(QWidget):
    def __init__(self, state_store: UiStateStore, paths: WorkoutPaths):
        super().__init__()
        self.state_store = state_store
        self.paths = paths

        self.workouts = []
        self.pinned = set()

        self._current_zoom = 1.0
        self._page_ready_pending = False

        self._today_primary_names = []
        self._today_alt_workouts = []
        self._today_phase = ""
        self._today_week = ""

        self._init_ui()
        self._install_web_scripts()

    def _init_ui(self):
        body = QHBoxLayout(self)
        body.setContentsMargins(14, 14, 14, 14)
        body.setSpacing(14)

        self.side_panel = QFrame()
        self.side_panel.setObjectName("SidePanel")
        side_lay = QVBoxLayout(self.side_panel)
        side_lay.setContentsMargins(12, 12, 12, 12)
        side_lay.setSpacing(10)

        self.search = QLineEdit()
        self.search.setPlaceholderText("FILTER WORKOUTS...")
        self.search.textChanged.connect(self._apply_filter)
        side_lay.addWidget(self.search)

        self.scroll = QScrollArea()
        self.scroll.setObjectName("WorkoutScroll")
        self.scroll.setFixedWidth(360)
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.scroll.verticalScrollBar().setStyleSheet(SCROLLBAR_STYLE)
        self.scroll.horizontalScrollBar().setStyleSheet(SCROLLBAR_STYLE)

        self.card_host = QWidget()
        self.card_host.setStyleSheet("background: transparent;")
        self.flow = FlowLayout(self.card_host)
        self.card_host.setLayout(self.flow)

        self.scroll.setWidget(self.card_host)
        self.scroll.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        side_lay.addWidget(self.scroll, stretch=1)

        body.addWidget(self.side_panel)

        self.web = QWebEngineView()
        self.web.setStyleSheet(f"background: #ffffff; border-radius: {RADIUS}px;")

        s = self.web.settings()
        s.setAttribute(QWebEngineSettings.WebAttribute.JavascriptEnabled, True)
        s.setAttribute(QWebEngineSettings.WebAttribute.LocalContentCanAccessFileUrls, True)
        s.setAttribute(QWebEngineSettings.WebAttribute.LocalContentCanAccessRemoteUrls, True)

        self.page = ExternalizingPage(self.web)
        self.web.setPage(self.page)

        self.web.page().zoomFactorChanged.connect(self._on_zoom_detected)
        self.web.loadFinished.connect(self._on_web_load_finished)

        body.addWidget(self.web, stretch=1)

    def _install_web_scripts(self):
        script = QWebEngineScript()
        script.setName("InjectScrollbarCSS")
        script.setInjectionPoint(QWebEngineScript.InjectionPoint.DocumentReady)
        script.setWorldId(QWebEngineScript.ScriptWorldId.MainWorld)
        script.setRunsOnSubFrames(True)
        script.setSourceCode(
            f"""
            (function() {{
                try {{
                    if (document.getElementById('__qt_scrollbar_style__')) return;
                    const st = document.createElement('style');
                    st.id = '__qt_scrollbar_style__';
                    st.type = 'text/css';
                    st.appendChild(document.createTextNode({json.dumps(WEB_SCROLLBAR_CSS)}));
                    (document.head || document.documentElement).appendChild(st);
                }} catch(e) {{}}
            }})();
            """
        )
        self.web.page().scripts().insert(script)

    def set_zoom(self, z: float):
        self._current_zoom = float(z)
        self.web.setZoomFactor(self._current_zoom)

    def zoom(self) -> float:
        return float(self._current_zoom)

    def _on_zoom_detected(self, factor: float):
        self._current_zoom = float(factor)

    # -------------------------
    # TODAY BAR (rendered by MainWindow)
    # -------------------------
    def today_model(self):
        return {
            "primary_names": list(self._today_primary_names),
            "alt_workouts": list(self._today_alt_workouts),
            "phase": self._today_phase,
            "week": self._today_week,
        }

    # -------------------------
    # WORKOUT LIST + PINS
    # -------------------------
    def clear_favourites(self):
        if not self.pinned:
            return
        self.pinned.clear()
        self._apply_filter(self.search.text() if self.search.text() else "")

    def toggle_pin(self, filename: str):
        if filename in self.pinned:
            self.pinned.remove(filename)
        else:
            self.pinned.add(filename)
        self._apply_filter(self.search.text() if self.search.text() else "")

    def reload_workouts(self, autoload=True):
        self.workouts = []
        if self.paths.workout_dir.exists():
            files = sorted(list(self.paths.workout_dir.glob("*.html")), key=lambda x: x.name.lower())
            for p in files:
                title = p.stem.replace("_", " ")
                self.workouts.append({"path": p, "title": title, "file": p.name})

        self._apply_filter(self.search.text() if self.search.text() else "")
        if autoload:
            self._autoload_default()

    def _autoload_default(self):
        if not self.workouts:
            return
        # pick first pinned (sorted) if any, else first alpha
        pinned_files = sorted([w for w in self.workouts if w["file"] in self.pinned], key=lambda w: w["title"].lower())
        target = pinned_files[0]["path"] if pinned_files else self.workouts[0]["path"]
        self.load_workout(target)

    def apply_filter_sort(self, text: str):
        t = (text or "").lower().strip()
        filtered = [w for w in self.workouts if t in w["title"].lower()]
        filtered.sort(key=lambda w: (w["file"] not in self.pinned, w["title"].lower()))
        return filtered

    def _apply_filter(self, text):
        while self.flow.count():
            item = self.flow.takeAt(0)
            if item and item.widget():
                item.widget().deleteLater()

        filtered = self.apply_filter_sort(text)

        for w in filtered:
            pinned = w["file"] in self.pinned
            card = WorkoutCard(
                w["title"],
                w["file"],
                pinned=pinned,
                on_open=lambda p=w["path"]: self.load_workout(p),
                on_pin_toggle=self.toggle_pin,
            )
            self.flow.addWidget(card)

        self.flow.invalidate()

    def load_workout(self, path: Path):
        self.web.setUrl(QUrl.fromLocalFile(str(path)))
        self.web.setZoomFactor(self._current_zoom)

    # -------------------------
    # WEB SCROLL RESTORE (state managed by MainWindow)
    # -------------------------
    def web_key(self) -> str:
        try:
            u = self.web.url()
            if u.isEmpty():
                return ""
            if u.isLocalFile():
                return "workout:" + Path(u.toLocalFile()).name
            return "workout:" + u.toString()
        except Exception:
            return ""

    def _on_web_load_finished(self, ok: bool):
        if not ok:
            return
        self.web.setZoomFactor(self._current_zoom)

        key = self.web_key()
        if not key:
            return

        st = self.state_store.load()
        web_state = st.get("web_state") or {}
        entry = web_state.get(key) or {}

        try:
            sx = int(entry.get("scroll_x", 0))
            sy = int(entry.get("scroll_y", 0))
        except Exception:
            sx, sy = 0, 0

        self._page_ready_pending = True
        QTimer.singleShot(120, lambda: self._restore_scroll(sx, sy))

    def _restore_scroll(self, sx: int, sy: int):
        if not self._page_ready_pending:
            return
        self._page_ready_pending = False
        self.web.page().runJavaScript(f"window.scrollTo({sx}, {sy});")

    def capture_scroll(self, cb):
        key = self.web_key()

        def _got_scroll(res):
            sx = int(res.get("x", 0)) if isinstance(res, dict) else 0
            sy = int(res.get("y", 0)) if isinstance(res, dict) else 0
            cb(key, sx, sy)

        self.web.page().runJavaScript("({x: window.scrollX, y: window.scrollY})", _got_scroll)

    # -------------------------
    # EXCEL TODAY LOAD
    # -------------------------
    def load_pins_from_excel_today(self):
        self.pinned = set()
        self._today_primary_names = []
        self._today_alt_workouts = []
        self._today_phase = ""
        self._today_week = ""

        if not self.paths.excel_path.exists():
            return

        try:
            wb = load_workbook(self.paths.excel_path, data_only=True)
            if "MASTER_CALENDAR" not in wb.sheetnames:
                return
            ws = wb["MASTER_CALENDAR"]

            header_row, header_map = None, {}
            for r in range(1, 21):
                vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                norm = [str(v).strip().upper() if v else "" for v in vals]
                if "PRIMARY CODE" in norm:
                    header_row = r
                    for c, v in enumerate(vals, start=1):
                        if v:
                            header_map[str(v).strip()] = c
                    break
            if not header_row:
                return

            today_s = date.today().strftime("%Y-%m-%d")
            match_row = None
            for r in range(header_row + 1, ws.max_row + 1):
                dval = ws.cell(row=r, column=1).value
                if isinstance(dval, (datetime, date)):
                    d_str = dval.strftime("%Y-%m-%d")
                else:
                    d_str = str(dval).strip()
                if d_str == today_s:
                    match_row = r
                    break
            if not match_row:
                return

            phase_col = header_map.get("Phase")
            if phase_col:
                pv = ws.cell(row=match_row, column=phase_col).value
                if pv:
                    self._today_phase = str(pv).strip()

            week_col = header_map.get("Week#")
            if week_col:
                wv = ws.cell(row=match_row, column=week_col).value
                if wv is not None:
                    self._today_week = str(wv).strip()

            for key in ["Primary Code", "Add-on 1", "Add-on 2", "Optional Add-on (SBX)"]:
                col = header_map.get(key)
                if col:
                    v = ws.cell(row=match_row, column=col).value
                    if v:
                        fn = str(v).strip()
                        if not fn.lower().endswith(".html"):
                            fn += ".html"
                        self.pinned.add(fn)
                        if key in ("Primary Code", "Optional Add-on (SBX)"):
                            self._today_primary_names.append(fn)

            for field_name, short_label in ALT_FIELD_MAP.items():
                col = header_map.get(field_name)
                if col:
                    v = ws.cell(row=match_row, column=col).value
                    if v:
                        fn = str(v).strip()
                        if not fn.lower().endswith(".html"):
                            fn += ".html"
                        self._today_alt_workouts.append((short_label, fn))
        except Exception:
            pass


class RegimenView(QWidget):
    def __init__(self, state_store: UiStateStore, paths: RegimenPaths):
        super().__init__()
        self.state_store = state_store
        self.paths = paths

        self._current_zoom = 1.0
        self._page_ready_pending = False

        self._init_ui()
        self._install_web_scripts()

    def _init_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(14, 14, 14, 14)
        lay.setSpacing(14)

        self.web = QWebEngineView()
        self.web.setStyleSheet(f"background: #ffffff; border-radius: {RADIUS}px;")

        s = self.web.settings()
        s.setAttribute(QWebEngineSettings.WebAttribute.JavascriptEnabled, True)
        s.setAttribute(QWebEngineSettings.WebAttribute.LocalContentCanAccessFileUrls, True)
        s.setAttribute(QWebEngineSettings.WebAttribute.LocalContentCanAccessRemoteUrls, True)

        self.page = ExternalizingPage(self.web)
        self.web.setPage(self.page)

        self.web.loadFinished.connect(self._on_web_load_finished)
        self.web.page().zoomFactorChanged.connect(self._on_zoom_detected)

        lay.addWidget(self.web, stretch=1)

        self.err = QLabel("", self)
        self.err.setObjectName("ErrorOverlay")
        self.err.setWordWrap(True)
        self.err.hide()

    def _install_web_scripts(self):
        script = QWebEngineScript()
        script.setName("InjectScrollbarCSS")
        script.setInjectionPoint(QWebEngineScript.InjectionPoint.DocumentReady)
        script.setWorldId(QWebEngineScript.ScriptWorldId.MainWorld)
        script.setRunsOnSubFrames(True)
        script.setSourceCode(
            f"""
            (function() {{
                try {{
                    if (document.getElementById('__qt_scrollbar_style__')) return;
                    const st = document.createElement('style');
                    st.id = '__qt_scrollbar_style__';
                    st.type = 'text/css';
                    st.appendChild(document.createTextNode({json.dumps(WEB_SCROLLBAR_CSS)}));
                    (document.head || document.documentElement).appendChild(st);
                }} catch(e) {{}}
            }})();
            """
        )
        self.web.page().scripts().insert(script)

    def set_zoom(self, z: float):
        self._current_zoom = float(z)
        self.web.setZoomFactor(self._current_zoom)

    def zoom(self) -> float:
        return float(self._current_zoom)

    def _on_zoom_detected(self, factor: float):
        self._current_zoom = float(factor)

    def web_key(self) -> str:
        try:
            u = self.web.url()
            if u.isEmpty():
                return ""
            if u.isLocalFile():
                return "regimen:" + Path(u.toLocalFile()).name
            return "regimen:" + u.toString()
        except Exception:
            return ""

    def _on_web_load_finished(self, ok: bool):
        if not ok:
            self._show_error(f"WEB LOAD FAILED:\n{self.web.url().toString()}")
            return

        self._hide_error()
        self.web.setZoomFactor(self._current_zoom)

        key = self.web_key()
        if not key:
            return

        st = self.state_store.load()
        web_state = st.get("web_state") or {}
        entry = web_state.get(key) or {}

        try:
            sx = int(entry.get("scroll_x", 0))
            sy = int(entry.get("scroll_y", 0))
        except Exception:
            sx, sy = 0, 0

        self._page_ready_pending = True
        QTimer.singleShot(120, lambda: self._restore_scroll(sx, sy))

    def _restore_scroll(self, sx: int, sy: int):
        if not self._page_ready_pending:
            return
        self._page_ready_pending = False
        self.web.page().runJavaScript(f"window.scrollTo({sx}, {sy});")

    def capture_scroll(self, cb):
        key = self.web_key()

        def _got_scroll(res):
            sx = int(res.get("x", 0)) if isinstance(res, dict) else 0
            sy = int(res.get("y", 0)) if isinstance(res, dict) else 0
            cb(key, sx, sy)

        self.web.page().runJavaScript("({x: window.scrollX, y: window.scrollY})", _got_scroll)

    def load_site(self):
        if not self.paths.html_path.exists():
            self._show_error(f"INDEX.HTML NOT FOUND:\n{self.paths.html_path}")
            return
        self._hide_error()
        self.web.setUrl(QUrl.fromLocalFile(str(self.paths.html_path)))
        self.web.setZoomFactor(self._current_zoom)

    def _show_error(self, msg: str):
        self.err.setText(msg)
        self.err.show()
        self.err.raise_()

    def _hide_error(self):
        self.err.hide()


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.setObjectName("Root")
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self.setWindowOpacity(WINDOW_OPACITY)

        self.state_store = UiStateStore()

        self._drag_pos = QPoint()
        self._dragging = False

        self._save_debounce = QTimer(self)
        self._save_debounce.setSingleShot(True)
        self._save_debounce.timeout.connect(self._save_state)

        self.base = self._resolve_base()
        self.workout_paths = self._resolve_workout_paths()
        self.regimen_paths = self._resolve_regimen_paths()

        self._init_icons()
        self._init_ui()
        self.setStyleSheet(build_theme_qss())

        self._load_state()

        # init data and views
        self.workout_view.load_pins_from_excel_today()
        self.workout_view.reload_workouts(autoload=True)
        self._refresh_header_today_bar()

        self.regimen_view.load_site()

        self._set_active_page(0)

    # -------------------------
    # PATH RESOLUTION (future-proof)
    # -------------------------
    def _resolve_base(self) -> Path:
        # Standalone mode: base is the directory containing this script
        return Path(__file__).resolve().parent

    def _resolve_workout_paths(self) -> WorkoutPaths:
        w_dir = self.base / WORKOUT_SUBPATH / WORKOUT_HTML_DIR
        e_path = self.base / WORKOUT_SUBPATH / WORKOUT_EXCEL
        return WorkoutPaths(base=self.base, workout_dir=w_dir, excel_path=e_path)

    def _resolve_regimen_paths(self) -> RegimenPaths:
        cand = self.base / REGIMEN_SUBPATH
        html_path = cand / REGIMEN_INDEX
        excel_path = cand / REGIMEN_EXCEL

        if html_path.exists():
            return RegimenPaths(site_dir=cand, html_path=html_path, excel_path=excel_path)

        # fallback: search within extensions for index.html
        root = self.base / "Main/gui/extensions"
        if root.exists():
            for p in [root] + list(root.rglob("*")):
                if p.is_dir() and (p / REGIMEN_INDEX).exists():
                    return RegimenPaths(site_dir=p, html_path=p / REGIMEN_INDEX, excel_path=p / REGIMEN_EXCEL)

        return RegimenPaths(site_dir=cand, html_path=html_path, excel_path=excel_path)

    def _init_icons(self):
        self.health_icon_path = self.base / HEALTH_ICON_REL
        self.alt_icon_path = self.base / ALT_ICON_REL

        if self.health_icon_path.exists():
            self.setWindowIcon(QIcon(str(self.health_icon_path)))
        elif self.alt_icon_path.exists():
            self.setWindowIcon(QIcon(str(self.alt_icon_path)))

    # -------------------------
    # UI
    # -------------------------
    def _init_ui(self):
        self.resize(*START_SIZE)

        root_lay = QVBoxLayout(self)
        root_lay.setContentsMargins(0, 0, 0, 0)

        self.container = QFrame()
        self.container.setObjectName("MainContainer")
        root_lay.addWidget(self.container)

        v_lay = QVBoxLayout(self.container)
        v_lay.setContentsMargins(0, 0, 0, 0)
        v_lay.setSpacing(0)

        # Header (single row - all buttons in one line)
        header = QFrame()
        header.setObjectName("Header")
        self.header = header

        row = QHBoxLayout(header)
        row.setContentsMargins(14, 10, 12, 10)
        row.setSpacing(10)

        # top-left health logo
        self.logo_lbl = QLabel()
        self.logo_lbl.setFixedSize(28, 28)
        self.logo_lbl.setScaledContents(True)
        if self.health_icon_path.exists():
            self.logo_lbl.setPixmap(QPixmap(str(self.health_icon_path)))
        elif self.alt_icon_path.exists():
            self.logo_lbl.setPixmap(QPixmap(str(self.alt_icon_path)))
        row.addWidget(self.logo_lbl)

        title_lbl = QLabel(APP_TITLE)
        title_lbl.setObjectName("Title")
        row.addWidget(title_lbl)

        self.phase_lbl = QLabel("")
        self.phase_lbl.setObjectName("PhaseLabel")
        row.addWidget(self.phase_lbl)

        # nav buttons
        self.btn_nav_workouts = QPushButton("WORKOUT PLANS")
        self.btn_nav_workouts.setObjectName("NavBtn")
        self.btn_nav_workouts.clicked.connect(lambda: self._set_active_page(0))
        row.addWidget(self.btn_nav_workouts)

        self.btn_nav_external = QPushButton("EXTERNAL APP")
        self.btn_nav_external.setObjectName("NavBtn")
        self.btn_nav_external.clicked.connect(lambda: self._set_active_page(1))
        row.addWidget(self.btn_nav_external)

        # spacer
        row.addStretch()

        # today bar buttons (center-right, still on same line)
        self.today_primary_layout = QHBoxLayout()
        self.today_primary_layout.setContentsMargins(0, 0, 0, 0)
        self.today_primary_layout.setSpacing(6)
        row.addLayout(self.today_primary_layout)

        self.today_sep_label = QLabel("\u2502")
        self.today_sep_label.setObjectName("TodaySep")
        self.today_sep_label.setFixedWidth(12)
        row.addWidget(self.today_sep_label)

        self.today_alt_layout = QHBoxLayout()
        self.today_alt_layout.setContentsMargins(0, 0, 0, 0)
        self.today_alt_layout.setSpacing(6)
        row.addLayout(self.today_alt_layout)

        row.addSpacing(14)

        # action buttons (workouts-aware; disabled on external page where not applicable)
        self.btn_clear_favs = QPushButton("DESELECT FAVS")
        self.btn_clear_favs.clicked.connect(self._clear_favourites)
        row.addWidget(self.btn_clear_favs)

        self.btn_excel = QPushButton("MASTER PLAN")
        self.btn_excel.clicked.connect(self._open_master_plan)
        row.addWidget(self.btn_excel)

        self.btn_folder = QPushButton("OPEN FOLDER")
        self.btn_folder.clicked.connect(self._open_folder)
        row.addWidget(self.btn_folder)

        self.btn_yt = QPushButton("YOUTUBE")
        self.btn_yt.setObjectName("YTBtn")
        self.btn_yt.clicked.connect(lambda: webbrowser.open("https://www.youtube.com/"))
        row.addWidget(self.btn_yt)

        self.btn_yt_music = QPushButton("YT MUSIC")
        self.btn_yt_music.setObjectName("YTMusicBtn")
        self.btn_yt_music.clicked.connect(lambda: webbrowser.open("https://music.youtube.com/"))
        row.addWidget(self.btn_yt_music)

        self.btn_spot = QPushButton("SPOTIFY")
        self.btn_spot.setObjectName("MusicBtn")
        self.btn_spot.clicked.connect(self._open_spotify_app)
        row.addWidget(self.btn_spot)

        # small secondary icon near window controls
        self.mini_icon = QLabel()
        self.mini_icon.setFixedSize(18, 18)
        self.mini_icon.setScaledContents(True)
        if self.alt_icon_path.exists():
            self.mini_icon.setPixmap(QPixmap(str(self.alt_icon_path)))
        elif self.health_icon_path.exists():
            self.mini_icon.setPixmap(QPixmap(str(self.health_icon_path)))
        row.addWidget(self.mini_icon)

        self.btn_min = QPushButton("\u2013")
        self.btn_min.setObjectName("MinButton")
        self.btn_min.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_min.clicked.connect(self.showMinimized)
        row.addWidget(self.btn_min)

        self.btn_close = QPushButton("X")
        self.btn_close.setObjectName("CloseButton")
        self.btn_close.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_close.clicked.connect(self.close)
        row.addWidget(self.btn_close)

        v_lay.addWidget(header)

        # pages
        self.stack = QStackedWidget()
        v_lay.addWidget(self.stack, stretch=1)

        self.workout_view = WorkoutView(self.state_store, self.workout_paths)
        self.regimen_view = RegimenView(self.state_store, self.regimen_paths)

        self.stack.addWidget(self.workout_view)
        self.stack.addWidget(self.regimen_view)

        # footer grip
        footer = QHBoxLayout()
        footer.addStretch()
        grip = QSizeGrip(self)
        grip.setStyleSheet("width: 20px; height: 20px; background: transparent;")
        footer.addWidget(grip)
        v_lay.addLayout(footer)

        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)

    # -------------------------
    # NAV / HEADER STATE
    # -------------------------
    def _set_nav_active(self, btn: QPushButton, active: bool):
        btn.setProperty("active", "true" if active else "false")
        btn.style().unpolish(btn)
        btn.style().polish(btn)
        btn.update()

    def _set_active_page(self, index: int):
        self.stack.setCurrentIndex(index)

        is_workouts = (index == 0)
        self._set_nav_active(self.btn_nav_workouts, is_workouts)
        self._set_nav_active(self.btn_nav_external, not is_workouts)

        # enable/disable workout-only controls
        for w in [self.btn_clear_favs, self.btn_folder]:
            w.setEnabled(is_workouts)

        # MASTER PLAN: switch target based on page
        self._refresh_header_today_bar()
        self._queue_save()

    def _clear_layout(self, layout):
        while layout.count():
            item = layout.takeAt(0)
            if item and item.widget():
                item.widget().deleteLater()

    def _refresh_header_today_bar(self):
        self._clear_layout(self.today_primary_layout)
        self._clear_layout(self.today_alt_layout)

        if self.stack.currentIndex() != 0:
            self.phase_lbl.setText("")
            self.today_sep_label.setVisible(False)
            return

        model = self.workout_view.today_model()
        prim = model["primary_names"]
        alt = model["alt_workouts"]

        parts = []
        if model["week"]:
            parts.append(f"WEEK: {model['week']}")
        if model["phase"]:
            parts.append(f"PHASE: {model['phase'].upper()}")
        self.phase_lbl.setText("\u25b8  " + "  \u00b7  ".join(parts) if parts else "")

        has_primary = bool(prim)
        has_alt = bool(alt)

        if has_primary:
            for fn in prim:
                display = fn.replace(".html", "").replace("_", " ").upper()
                btn = QPushButton(display)
                btn.setObjectName("TodayPrimaryBtn")
                btn.setCursor(Qt.CursorShape.PointingHandCursor)
                btn.clicked.connect(lambda checked=False, f=fn: self._today_btn_clicked(f))
                self.today_primary_layout.addWidget(btn)
        else:
            lbl = QLabel("REST DAY")
            lbl.setObjectName("TodayNone")
            self.today_primary_layout.addWidget(lbl)

        self.today_sep_label.setVisible(has_primary and has_alt)

        if has_alt:
            for short_label, fn in alt:
                code_display = fn.replace(".html", "").replace("_", " ").upper()
                btn = QPushButton(f"{short_label}: {code_display}")
                btn.setObjectName("TodayAltBtn")
                btn.setCursor(Qt.CursorShape.PointingHandCursor)
                btn.clicked.connect(lambda checked=False, f=fn: self._today_btn_clicked(f))
                self.today_alt_layout.addWidget(btn)

    def _today_btn_clicked(self, filename: str):
        target = self.workout_paths.workout_dir / filename
        if target.exists():
            self.workout_view.load_workout(target)

    # -------------------------
    # BUTTON ACTIONS
    # -------------------------
    def _open_spotify_app(self):
        try:
            os.startfile("spotify:")
        except Exception:
            try:
                subprocess.Popen(["explorer.exe", "spotify:"])
            except Exception:
                webbrowser.open("https://open.spotify.com/")

    def _open_folder(self):
        if self.workout_paths.workout_dir.exists():
            os.startfile(self.workout_paths.workout_dir)

    def _open_master_plan(self):
        if self.stack.currentIndex() == 0:
            p = self.workout_paths.excel_path
        else:
            p = self.regimen_paths.excel_path
        if p.exists():
            os.startfile(p)

    def _clear_favourites(self):
        self.workout_view.clear_favourites()
        # keep today-bar pins intact? user asked deselect favs; do not touch excel-derived today pins
        # but keep UI consistent:
        self._refresh_header_today_bar()

    # -------------------------
    # CONTEXT MENU
    # -------------------------
    def _show_context_menu(self, pos):
        menu = QMenu(self)

        if self.stack.currentIndex() == 0:
            act_clear = QAction("Deselect favourites", self)
            act_clear.triggered.connect(self._clear_favourites)
            menu.addAction(act_clear)
            menu.addSeparator()

            act_excel = QAction("Open master plan", self)
            act_excel.triggered.connect(self._open_master_plan)
            menu.addAction(act_excel)

            act_folder = QAction("Open folder", self)
            act_folder.triggered.connect(self._open_folder)
            menu.addAction(act_folder)
        else:
            act_excel = QAction("Open master plan", self)
            act_excel.triggered.connect(self._open_master_plan)
            menu.addAction(act_excel)

            act_reload = QAction("Reload", self)
            act_reload.triggered.connect(self.regimen_view.load_site)
            menu.addAction(act_reload)

        menu.addSeparator()

        act_min = QAction("Minimize", self)
        act_min.triggered.connect(self.showMinimized)
        menu.addAction(act_min)

        act_close = QAction("Close", self)
        act_close.triggered.connect(self.close)
        menu.addAction(act_close)

        menu.exec(self.mapToGlobal(pos))

    # -------------------------
    # STATE SAVE / RESTORE
    # -------------------------
    def _load_state(self):
        st = self.state_store.load()
        val = st.get("universal_zoom")
        z = float(val) if val is not None else 1.0

        self.workout_view.set_zoom(z)
        self.regimen_view.set_zoom(z)

        win = st.get("window") or {}
        try:
            x, y = int(win.get("x", 100)), int(win.get("y", 100))
            w, h = int(win.get("w", START_SIZE[0])), int(win.get("h", START_SIZE[1]))
            if w > 200 and h > 200:
                self.setGeometry(x, y, w, h)
            else:
                self.resize(*START_SIZE)
        except Exception:
            self.resize(*START_SIZE)

    def _capture_scroll_and_save(self):
        st = self.state_store.load()
        st["universal_zoom"] = float(self.workout_view.zoom() if self.stack.currentIndex() == 0 else self.regimen_view.zoom())

        web_state = st.get("web_state") or {}

        pending = {"count": 0}

        def _done():
            pending["count"] -= 1
            if pending["count"] <= 0:
                g = self.geometry()
                st["window"] = {"x": g.x(), "y": g.y(), "w": g.width(), "h": g.height()}
                st["web_state"] = web_state
                st["active_page"] = int(self.stack.currentIndex())
                self.state_store.save(st)

        def _store(key, sx, sy):
            if key:
                entry = web_state.get(key) or {}
                entry["scroll_x"] = int(sx)
                entry["scroll_y"] = int(sy)
                web_state[key] = entry
            _done()

        # capture both views to keep restore consistent
        pending["count"] = 2
        self.workout_view.capture_scroll(_store)
        self.regimen_view.capture_scroll(_store)

    def _save_state(self):
        self._capture_scroll_and_save()

    def _queue_save(self):
        self._save_debounce.start(250)

    # -------------------------
    # WINDOW EVENTS
    # -------------------------
    def closeEvent(self, e):
        self._save_state()
        super().closeEvent(e)

    def moveEvent(self, e):
        super().moveEvent(e)
        self._queue_save()

    def resizeEvent(self, e):
        super().resizeEvent(e)
        self._queue_save()

    # -------------------------
    # DRAG HEADER
    # -------------------------
    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            click_pos = event.position().toPoint()
            header_rect = self.header.geometry()
            if header_rect.contains(click_pos):
                child = self.childAt(click_pos)
                if child is None or isinstance(child, (QFrame, QLabel)):
                    self._dragging = True
                    self._drag_pos = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
                    event.accept()
                    return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._dragging and event.buttons() == Qt.MouseButton.LeftButton:
            self.move(event.globalPosition().toPoint() - self._drag_pos)
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self._dragging = False
        super().mouseReleaseEvent(event)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    QWebEngineProfile.defaultProfile().setHttpCacheType(QWebEngineProfile.HttpCacheType.MemoryHttpCache)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())
