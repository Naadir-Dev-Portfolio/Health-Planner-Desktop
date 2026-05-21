#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import shutil
import sys
import time
from datetime import date, datetime, time as dtime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


def ts() -> float:
    return time.perf_counter()


def log(msg: str) -> None:
    print(msg, file=sys.stdout, flush=True)


def jsonable(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (datetime, date, dtime)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    return v


def is_good_header_row(row: Sequence[Any]) -> bool:
    vals = [v for v in row if v is not None and str(v).strip() != ""]
    if len(vals) < 2:
        return False
    strs = [str(v).strip() for v in vals]
    return len(set(strs)) == len(strs)


def normalize_headers(headers: Sequence[Any]) -> List[str]:
    used = set()
    out: List[str] = []
    for i, h in enumerate(headers, start=1):
        name = (str(h).strip() if h is not None else "") or f"col_{i}"
        base = name
        k = 2
        while name in used:
            name = f"{base}_{k}"
            k += 1
        used.add(name)
        out.append(name)
    return out


def trim_trailing_blanks(row: Sequence[Any]) -> Tuple[List[Any], int]:
    r = list(row)
    last = -1
    for i, v in enumerate(r):
        if v is not None and str(v).strip() != "":
            last = i
    if last == -1:
        return [], 0
    return r[: last + 1], last + 1


def sample_rows(ws, n: int) -> List[List[Any]]:
    out: List[List[Any]] = []
    it = ws.iter_rows(values_only=True)
    for _ in range(n):
        try:
            out.append(list(next(it)))
        except StopIteration:
            break
    return out


def find_header(sample: List[List[Any]]) -> Tuple[int, List[str], int]:
    if not sample:
        return 1, [], 0

    best_idx = 0
    best_score = -1
    best_headers: Optional[List[str]] = None
    best_width = 0

    for idx, row in enumerate(sample):
        trimmed, width = trim_trailing_blanks(row)
        if width == 0:
            continue
        if not is_good_header_row(trimmed):
            continue
        score = sum(1 for v in trimmed if v is not None and str(v).strip() != "")
        if score > best_score:
            best_score = score
            best_idx = idx
            best_headers = normalize_headers(trimmed)
            best_width = width

    if best_headers is None:
        trimmed, width = trim_trailing_blanks(sample[0])
        width = width or (len(sample[0]) if sample[0] else 1)
        best_headers = [f"col_{i}" for i in range(1, width + 1)]
        best_idx = 0
        best_width = width

    return best_idx + 1, best_headers, best_width  # 1-based row number


def export_single_sheet(
    xlsm_path: Path,
    sheet_name: str,
    out_js_path: Path,
    header_scan_rows: int = 50,
    empty_row_stop: int = 100,
    progress_every: int = 5000,
    data_only: bool = True,
) -> Dict[str, Any]:
    t0 = ts()
    log(f"[1/5] Opening workbook: {xlsm_path.name}")
    t_open = ts()
    wb = load_workbook(
        filename=str(xlsm_path),
        read_only=True,
        data_only=data_only,
        keep_vba=True,
    )
    log(f"      -> opened in {(ts() - t_open):.3f}s | sheets={len(wb.sheetnames)}")

    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {sheet_name!r}. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    log(f"[2/5] Scanning header (first {header_scan_rows} rows) on sheet: {sheet_name!r}")
    t_scan = ts()
    sample = sample_rows(ws, header_scan_rows)
    header_row, headers, width = find_header(sample)
    log(
        f"      -> header_row={header_row} | columns={len(headers)} | width={width} | scan_time={(ts() - t_scan):.3f}s"
    )

    log(f"[3/5] Exporting records (stops after {empty_row_stop} consecutive empty rows)")
    t_iter = ts()
    it = ws.iter_rows(min_row=header_row + 1, values_only=True)

    records: List[Dict[str, Any]] = []
    empty_streak = 0
    seen_rows = 0

    for row in it:
        seen_rows += 1
        trimmed = list(row[:width]) if width > 0 else list(row)

        if all(v is None or str(v).strip() == "" for v in trimmed):
            empty_streak += 1
            if empty_streak >= empty_row_stop:
                log(f"      -> stop: hit empty_row_stop at seen_rows={seen_rows} | records={len(records)}")
                break
            continue

        empty_streak = 0
        rec: Dict[str, Any] = {}
        for i, key in enumerate(headers):
            rec[key] = jsonable(trimmed[i] if i < len(trimmed) else None)
        records.append(rec)

        if progress_every and (len(records) % progress_every == 0):
            log(f"      -> progress: records={len(records)} | seen_rows={seen_rows}")

    log(f"      -> exported records={len(records)} | iter_time={(ts() - t_iter):.3f}s")

    payload: Dict[str, Any] = {
        "name": ws.title,
        "header_row": header_row,
        "headers": headers,
        "row_count": len(records),
        "records": records,
    }

    log(f"[4/5] Writing JS bundle: {out_js_path}")
    t_write = ts()
    out_js_path.parent.mkdir(parents=True, exist_ok=True)
    with out_js_path.open("w", encoding="utf-8") as f:
        f.write("window.__ROUTINE_PAYLOAD__ = ")
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")
    size = out_js_path.stat().st_size
    log(f"      -> wrote {size:,} bytes in {(ts() - t_write):.3f}s")

    log(f"[5/5] Done in {(ts() - t0):.3f}s")
    return payload


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "input",
        nargs="?",
        default=None,
        help="Path to .xlsm/.xlsx (default: master_plan.xlsm next to this script)",
    )
    ap.add_argument(
        "--sheet",
        default="Routine",
        help="Sheet name to export (default: Routine)",
    )
    ap.add_argument(
        "--out",
        default="data/routine.data.js",
        help="Output JS path relative to script directory (default: data/routine.data.js)",
    )
    ap.add_argument("--header-scan-rows", type=int, default=50)
    ap.add_argument("--empty-row-stop", type=int, default=100)
    ap.add_argument("--progress-every", type=int, default=5000)
    ap.add_argument("--data-only", action="store_true", default=True)
    ap.add_argument(
        "--reset-data-folder",
        action="store_true",
        help="If set, deletes the entire ./data folder before writing output",
    )
    args = ap.parse_args()

    script_dir = Path(__file__).resolve().parent
    xlsm_path = (Path(args.input).resolve() if args.input else (script_dir / "master_plan.xlsm").resolve())

    out_path = (script_dir / args.out).resolve()
    data_dir = (script_dir / "data").resolve()

    if out_path.suffix.lower() != ".js":
        raise SystemExit(f"--out must be a .js file path, got: {out_path}")

    if args.reset_data_folder:
        log(f"[prep] Resetting folder: {data_dir}")
        if data_dir.exists():
            shutil.rmtree(data_dir)

    export_single_sheet(
        xlsm_path=xlsm_path,
        sheet_name=args.sheet,
        out_js_path=out_path,
        header_scan_rows=args.header_scan_rows,
        empty_row_stop=args.empty_row_stop,
        progress_every=args.progress_every,
        data_only=bool(args.data_only),
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
